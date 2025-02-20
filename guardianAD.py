import sys
import re
import argparse
import codecs
import datetime
from prettytable import PrettyTable
import html
from openpyxl import Workbook
from collections import defaultdict

def process_files(hashcat_file, ntds_file, filter_history=False, filter_computers=False, only_enabled=False):
    cracked_passwords = {}
    users_with_cracked_passwords = []
    password_usage = defaultdict(list)
    hash_usage = defaultdict(list)

    stats = {
        'total_accounts': 0,
        'total_user_accounts': 0,
        'total_computer_accounts': 0,
        'enabled_user_accounts': 0,
        'enabled_computer_accounts': 0,
        'cracked_accounts': 0,
        'cracked_enabled_accounts': 0
    }

    with open(hashcat_file, 'r') as f:
        for line in f:
            match = re.match(r'((?:[^\\]+\\)?[^:]+):([^:]+):(.+)', line.strip())
            if match:
                domain_username, ntlm_hash, password = match.groups()
                if password.startswith('$HEX[') and password.endswith(']'):
                    hex_content = password[5:-1]
                    password = codecs.decode(hex_content, 'hex').decode('ascii', errors='ignore')
                cracked_passwords[ntlm_hash] = password

    with open(ntds_file, 'r') as f:
        for line in f:
            match = re.match(r'((?:[^\\]+\\)?[^:]+):([^:]+):([^:]+):([^:]+):::(?:\s*\(status=([^)]+)\))?', line.strip())
            if match:
                domain_username, user_id, lm_hash, ntlm_hash, status = match.groups()
                is_history = '_history' in domain_username.lower()
                is_computer = '$' in domain_username
                is_enabled = status and status.lower() == 'enabled'

                # Apply filters first
                if any([
                    filter_history and is_history,
                    filter_computers and is_computer,
                    only_enabled and not is_enabled
                ]):
                    continue

                # Update statistics for non-filtered accounts
                stats['total_accounts'] += 1
                if is_computer:
                    stats['total_computer_accounts'] += 1
                    if is_enabled:
                        stats['enabled_computer_accounts'] += 1
                else:
                    stats['total_user_accounts'] += 1
                    if is_enabled:
                        stats['enabled_user_accounts'] += 1

                hash_usage[ntlm_hash].append(domain_username)

                if ntlm_hash in cracked_passwords:
                    password = cracked_passwords[ntlm_hash]
                    users_with_cracked_passwords.append((domain_username, password))
                    password_usage[password].append(domain_username)
                    stats['cracked_accounts'] += 1
                    if is_enabled:
                        stats['cracked_enabled_accounts'] += 1

    results = []
    for domain_username, password in users_with_cracked_passwords:
        password_length = len(password)
        reused = "Yes" if len(password_usage[password]) > 1 else "No"
        results.append({
            "User": domain_username,
            "Password Length": password_length,
            "Password Preview": password[:6] + '*' * (password_length - 6),
            "Password Reused": reused,
            "Reused By": ", ".join(u for u in password_usage[password] if u != domain_username) if reused == "Yes" else "N/A"
        })

    return results, stats, users_with_cracked_passwords, hash_usage

def calculate_similarity(str1, str2):
    set1, set2 = set(str1), set(str2)
    intersection = set1 & set2
    union = set1 | set2
    return len(intersection)/len(union)*100 if union else 0

def find_similar_passwords(users_with_cracked_passwords, threshold=70):
    similar = []
    for i, (u1, p1) in enumerate(users_with_cracked_passwords):
        for j, (u2, p2) in enumerate(users_with_cracked_passwords[i+1:], i+1):
            similarity = calculate_similarity(p1, p2)
            if 70 <= similarity < 100:
                masked = lambda p: p[:6] + '*'*(len(p)-6)
                similar.append((u1, u2, masked(p1), masked(p2), similarity))
    return similar

def create_ascii_table(results):
    table = PrettyTable()
    table.field_names = ["User", "Len", "Preview", "Reused", "Reused By"]
    table.align["User"] = "l"
    table.align["Preview"] = "l"
    table.align["Reused By"] = "l"
    table.max_width["User"] = 30
    table.max_width["Preview"] = 20
    table.max_width["Reused By"] = 30

    for row in results:
        table.add_row([
            row["User"][:30],
            row["Password Length"],
            row["Password Preview"][:20],
            "Y" if row["Password Reused"] == "Yes" else "N",
            (row["Reused By"] if row["Reused By"] != "N/A" else "")[:30]
        ])
    return table.get_string()

def create_html_table(results, stats, args, similar_passwords, hash_usage):
    current_date = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    filters = []
    if args.filter_history:
        filters.append("History filtered")
    if args.filter_computers:
        filters.append("Computers filtered")
    if args.only_enabled:
        filters.append("Only enabled")

    html_content = f"""
    <!DOCTYPE html>
    <html lang="en">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Password Security Report</title>
        <style>
            body {{
                font-family: 'Segoe UI', Arial, sans-serif;
                line-height: 1.6;
                color: #333;
                background: #f8f9fa;
                margin: 0;
                padding: 20px;
            }}
            .container {{
                max-width: 1200px;
                margin: 0 auto;
                background: white;
                padding: 30px;
                border-radius: 8px;
                box-shadow: 0 2px 15px rgba(0,0,0,0.1);
            }}
            h1, h2 {{
                color: #2c3e50;
                margin-bottom: 1.5rem;
            }}
            .stats-grid {{
                display: grid;
                grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
                gap: 1rem;
                margin-bottom: 2rem;
            }}
            .stat-card {{
                background: #3498db;
                color: white;
                padding: 1rem;
                border-radius: 5px;
                text-align: center;
            }}
            table {{
                width: 100%;
                border-collapse: collapse;
                margin: 1rem 0;
                box-shadow: 0 1px 3px rgba(0,0,0,0.1);
            }}
            th, td {{
                padding: 12px 15px;
                text-align: left;
                border-bottom: 1px solid #ddd;
            }}
            th {{
                background-color: #2c3e50;
                color: white;
                cursor: pointer;
            }}
            th:hover {{
                background-color: #34495e;
            }}
            tr:nth-child(even) {{
                background-color: #f9f9f9;
            }}
        </style>
        <script>
            function sortTable(tableId, columnIndex) {{
                var table, rows, switching, i, x, y, shouldSwitch, dir, switchcount = 0;
                table = document.getElementById(tableId);
                switching = true;
                dir = "asc";
                while (switching) {{
                    switching = false;
                    rows = table.rows;
                    for (i = 1; i < (rows.length - 1); i++) {{
                        shouldSwitch = false;
                        x = rows[i].getElementsByTagName("TD")[columnIndex];
                        y = rows[i + 1].getElementsByTagName("TD")[columnIndex];
                        if (dir == "asc") {{
                            if (isNaN(x.innerHTML) && isNaN(y.innerHTML)) {{
                                if (x.innerHTML.toLowerCase() > y.innerHTML.toLowerCase()) {{
                                    shouldSwitch = true;
                                    break;
                                }}
                            }} else {{
                                if (Number(x.innerHTML) > Number(y.innerHTML)) {{
                                    shouldSwitch = true;
                                    break;
                                }}
                            }}
                        }} else if (dir == "desc") {{
                            if (isNaN(x.innerHTML) && isNaN(y.innerHTML)) {{
                                if (x.innerHTML.toLowerCase() < y.innerHTML.toLowerCase()) {{
                                    shouldSwitch = true;
                                    break;
                                }}
                            }} else {{
                                if (Number(x.innerHTML) < Number(y.innerHTML)) {{
                                    shouldSwitch = true;
                                    break;
                                }}
                            }}
                        }}
                    }}
                    if (shouldSwitch) {{
                        rows[i].parentNode.insertBefore(rows[i + 1], rows[i]);
                        switching = true;
                        switchcount++;
                    }} else {{
                        if (switchcount == 0 && dir == "asc") {{
                            dir = "desc";
                            switching = true;
                        }}
                    }}
                }}
            }}
        </script>
    </head>
    <body>
        <div class="container">
            <h1>Password Security Analysis Report</h1>

            <div class="stats-grid">
                <div class="stat-card">
                    <h3>Total Accounts</h3>
                    <p>{stats['total_accounts']}</p>
                </div>
                <div class="stat-card">
                    <h3>Cracked Accounts</h3>
                    <p>{stats['cracked_accounts']} ({stats['cracked_accounts']/(stats['total_accounts'] if stats['total_accounts'] != 0 else 1)*100:.1f}%)</p>
                </div>
                <div class="stat-card">
                    <h3>Enabled Users</h3>
                    <p>{stats['enabled_user_accounts']}</p>
                </div>
                <div class="stat-card">
                    <h3>Reused Passwords</h3>
                    <p>{sum(1 for row in results if row['Password Reused'] == 'Yes')}</p>
                </div>
            </div>

            <h2>Cracked Passwords</h2>
            <table id="crackedTable">
                <thead>
                    <tr>
                        <th onclick="sortTable('crackedTable', 0)">User</th>
                        <th onclick="sortTable('crackedTable', 1)">Length</th>
                        <th onclick="sortTable('crackedTable', 2)">Preview</th>
                        <th onclick="sortTable('crackedTable', 3)">Reused</th>
                        <th onclick="sortTable('crackedTable', 4)">Reused By</th>
                    </tr>
                </thead>
                <tbody>
                    {"".join(f'''
                    <tr>
                        <td>{html.escape(row['User'])}</td>
                        <td>{row['Password Length']}</td>
                        <td>{html.escape(row['Password Preview'])}</td>
                        <td>{row['Password Reused']}</td>
                        <td>{html.escape(row['Reused By'])}</td>
                    </tr>
                    ''' for row in results)}
                </tbody>
            </table>

            <h2>Similar Passwords</h2>
            <table id="similarTable">
                <thead>
                    <tr>
                        <th onclick="sortTable('similarTable', 0)">User 1</th>
                        <th onclick="sortTable('similarTable', 1)">User 2</th>
                        <th onclick="sortTable('similarTable', 2)">Password 1</th>
                        <th onclick="sortTable('similarTable', 3)">Password 2</th>
                        <th onclick="sortTable('similarTable', 4)">Similarity</th>
                    </tr>
                </thead>
                <tbody>
                    {"".join(f'''
                    <tr>
                        <td>{html.escape(u1)}</td>
                        <td>{html.escape(u2)}</td>
                        <td>{html.escape(p1)}</td>
                        <td>{html.escape(p2)}</td>
                        <td>{sim:.1f}</td>
                    </tr>
                    ''' for u1, u2, p1, p2, sim in similar_passwords)}
                </tbody>
            </table>

            <h2>Shared Hashes</h2>
            <table id="sharedTable">
                <thead>
                    <tr>
                        <th onclick="sortTable('sharedTable', 0)">Hash</th>
                        <th onclick="sortTable('sharedTable', 1)">Users</th>
                    </tr>
                </thead>
                <tbody>
                    {"".join(f'''
                    <tr>
                        <td>{html.escape(hash_value[:-16] + '*' * 16 if len(hash_value) > 16 else '*' * 32)}</td>
                        <td>{html.escape(', '.join(users))}</td>
                    </tr>
                    ''' for hash_value, users in hash_usage.items() if len(users) > 1)}
                </tbody>
            </table>

            <div style="margin-top: 2rem; color: #666; font-size: 0.9em;">
                <p>Generated on: {current_date}</p>
                <p>Filters applied: {', '.join(filters) or 'None'}</p>
            </div>
        </div>
    </body>
    </html>
    """
    return html_content

def show_banner():
    print("""
 ██████╗ ██╗   ██╗ █████╗ ██████╗ ██████╗ ██╗ █████╗ ███╗   ██╗ █████╗ ██████╗
██╔════╝ ██║   ██║██╔══██╗██╔══██╗██╔══██╗██║██╔══██╗████╗  ██║██╔══██╗██╔══██╗
██║  ███╗██║   ██║███████║██████╔╝██║  ██║██║███████║██╔██╗ ██║███████║██║  ██║
██║   ██║██║   ██║██╔══██║██╔══██╗██║  ██║██║██╔══██║██║╚██╗██║██╔══██║██║  ██║
╚██████╔╝╚██████╔╝██║  ██║██║  ██║██████╔╝██║██║  ██║██║ ╚████║██║  ██║██████╔╝
 ╚═════╝  ╚═════╝ ╚═╝  ╚═╝╚═╝  ╚═╝╚═════╝ ╚═╝╚═╝  ╚═╝╚═╝  ╚═══╝╚═╝  ╚═╝╚═════╝
╔═════════════════════════════════════════════════════════════════════════════╗
║ 🛡️ Active Directory Password Security Dashboard v1.0                         ║
║ ▶ Hashcat Analyzer ◼ NTDS Inspector ◼ Password Strength Auditor ◼           ║
╚═════════════════════════════════════════════════════════════════════════════╝

    """)


def main():
    show_banner()
    parser = argparse.ArgumentParser(description="Analyze password hashes from Hashcat output and NTDS dump")
    parser.add_argument("hashcat_file", help="Path to Hashcat output file")
    parser.add_argument("ntds_file", help="Path to NTDS dump file")
    parser.add_argument("--filter-history", action="store_true", help="Filter out password history entries")
    parser.add_argument("--filter-computers", action="store_true", help="Filter out computer accounts")
    parser.add_argument("--only-enabled", action="store_true", help="Only include enabled accounts")
    parser.add_argument("--output", choices=["ascii", "html", "excel"], default="ascii", help="Output format")
    parser.add_argument("--output-file", help="Output file name (without extension)")
    args = parser.parse_args()

    results, stats, users_with_cracked_passwords, hash_usage = process_files(
        args.hashcat_file, args.ntds_file,
        filter_history=args.filter_history,
        filter_computers=args.filter_computers,
        only_enabled=args.only_enabled
    )

    similar_passwords = find_similar_passwords(users_with_cracked_passwords)

    if args.output == "ascii":
        print(create_ascii_table(results))
    elif args.output == "html":
        html_content = create_html_table(results, stats, args, similar_passwords, hash_usage)
        output_file = args.output_file if args.output_file else "password_analysis_report"
        file_name = f"{output_file}.html"
        with open(file_name, "w") as f:
            f.write(html_content)
        print(f"HTML report saved as {file_name}")
    elif args.output == "excel":
        wb = Workbook()

        # Cracked Passwords sheet
        ws = wb.active
        ws.title = "Cracked Passwords"
        ws.append(["User", "Password Length", "Password Preview", "Password Reused", "Reused By"])
        for row in results:
            ws.append([row["User"], row["Password Length"], row["Password Preview"], row["Password Reused"], row["Reused By"]])

        # Statistics sheet
        ws_stats = wb.create_sheet("Statistics")
        ws_stats.append(["Metric", "Value"])
        ws_stats.append(["Total accounts", stats['total_accounts']])
        ws_stats.append(["User accounts", stats['total_user_accounts']])
        ws_stats.append(["Enabled user accounts", stats['enabled_user_accounts']])
        ws_stats.append(["Computer accounts", stats['total_computer_accounts']])
        ws_stats.append(["Enabled computer accounts", stats['enabled_computer_accounts']])
        ws_stats.append(["Cracked accounts", stats['cracked_accounts']])
        ws_stats.append(["Cracked accounts percentage", f"{stats['cracked_accounts']/stats['total_accounts']*100:.2f}%"])
        ws_stats.append(["Cracked enabled accounts", stats['cracked_enabled_accounts']])
        ws_stats.append(["Cracked enabled accounts percentage", f"{stats['cracked_enabled_accounts']/stats['enabled_user_accounts']*100:.2f}%"])

        # Similar Passwords sheet
        ws_similar = wb.create_sheet("Similar Passwords")
        ws_similar.append(["User 1", "User 2", "Password 1", "Password 2", "Similarity"])
        for u1, u2, p1, p2, sim in similar_passwords:
            ws_similar.append([u1, u2, p1, p2, f"{sim:.1f}%"])

        # Shared Hashes sheet
        ws_shared = wb.create_sheet("Shared Hashes")
        ws_shared.append(["Hash", "Users"])
        for hash_value, users in hash_usage.items():
            if len(users) > 1:
                masked_hash = hash_value[:-16] + '*' * 16 if len(hash_value) > 16 else '*'*32
                ws_shared.append([masked_hash, ", ".join(users)])

        output_file = args.output_file if args.output_file else "password_analysis_report"
        file_name = f"{output_file}.xlsx"
        wb.save(file_name)
        print(f"Excel report saved as {file_name}")


if __name__ == "__main__":
    main()
